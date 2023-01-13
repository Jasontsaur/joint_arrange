''' 試安排樑柱接頭平面配置'''
import os
import sys
import math
import openpyxl     # for EXCEL access
import matplotlib.pyplot as plt
import matplotlib.lines as lines
# import check_rebar  # 排列梁柱接頭平面
import check_rebar_2  # 排列梁柱接頭平面


def line(x1,y1,x2,y2,c,lw):
    '''畫線 帶顏色'''
    fig.add_artist(lines.Line2D(wnd([x1,x2]), wnd([y1, y2]),color=c,linewidth=lw))

def draw_rebar(y,cmd:str):
    ''' 雙線表示鋼筋 '''
    color='#D0D0D0' #   灰色
    HDB=32/1100/2.0
    if cmd=='=':
        line(-0.1,y-HDB,1.1,y-HDB,color,2)
        line(-0.1,y+HDB,1.1,y+HDB,color,2)
    elif cmd=='L':  #   左端彎鉤
        line(0.1,y-HDB,1.1,y-HDB,color,2)
        line(0.1,y+HDB,1.1,y+HDB,color,2)
        draw_circle(0.1,y,HDB,color)
    elif cmd=='R':  #   右端彎鉤
        line(-0.1,y-HDB,0.9,y-HDB,color,2)
        line(-0.1,y+HDB,0.9,y+HDB,color,2)
        draw_circle(0.9,y,HDB,color)

def draw_circle(x,y,r,c):
    ''' 畫圓 用於柱主筋 與 表示梁筋的彎鉤 '''
    NSEG=8
    xlist=[]
    ylist=[]
    for _ in range(NSEG+1):
        theta=_*2*3.14159/NSEG
        xlist=xlist+[x+r*math.cos(theta)]
        ylist=ylist+[y+r*math.sin(theta)]
    fig.add_artist(lines.Line2D(wnd(xlist), wnd(ylist),color=c,linewidth=2))
    
def draw_beam_column_joint(_widc,_ba,_bb,_shift_a,_shift_b,_EDGE_BM):
    
     # 畫 柱子範圍

    fig.add_artist(lines.Line2D(wnd([0, 0]), wnd([0, _widc])))
    fig.add_artist(lines.Line2D(wnd([0, _widc]), wnd([_widc, _widc])))
    fig.add_artist(lines.Line2D(wnd([_widc, _widc]), wnd([_widc, 0])))
    fig.add_artist(lines.Line2D(wnd([_widc, 0]),wnd([0, 0])))
    # 畫大梁偏位置
    fig.add_artist(lines.Line2D(wnd([-0.1, 0]), wnd([_shift_a, _shift_a])))
    fig.add_artist(lines.Line2D(wnd([-0.1, 0]), wnd([(_ba+_shift_a), (_ba+_shift_a)])))
    fig.add_artist(lines.Line2D(wnd([1, 1.1]), wnd([_shift_b, _shift_b])))
    fig.add_artist(lines.Line2D(wnd([1, 1.1]), wnd([(_bb+_shift_b), (_bb+_shift_b)])))

    # 畫大梁鋼筋範圍 虛線 細線
    fig.add_artist(lines.Line2D(wnd([-0.1, 0]), wnd([(_shift_a+_EDGE_BM), (_shift_a+_EDGE_BM)]),\
        linestyle='--',linewidth=1))
    fig.add_artist(lines.Line2D(wnd([-0.1, 0]), wnd([(_ba+_shift_a-_EDGE_BM), (_ba+_shift_a-_EDGE_BM)]),\
        linestyle='--',linewidth=1))
    fig.add_artist(lines.Line2D(wnd([1, 1.1]), wnd([(_shift_b+_EDGE_BM), (_shift_b+_EDGE_BM)]),\
        linestyle='--',linewidth=1))
    fig.add_artist(lines.Line2D(wnd([1, 1.1]), wnd([(_bb+_shift_b-_EDGE_BM), (_bb+_shift_b-_EDGE_BM)]),\
        linestyle='--',linewidth=1))
    


def wnd(xlist):
    '''# 縮放視圖'''
    EDGE=0.1
    for _, element_i  in enumerate(xlist):
        xlist[_]=EDGE+element_i*(1-2*EDGE)
    return xlist

def getvalue(_element):
    ''' get symbol of a list'''
    return _element['value']


def select_runway(_n_ab,demand,y,occupied,out1,y_min,y_max,R_VALUE):
    '''尋求以盡量均佈方式,在y中挑選所需要的demand個位置,擺放梁筋'''
    # _n_runway:可以放梁筋的跑道數目
    # demand:需要挑選的鋼筋數
    # y(_n_runway):空格座標
    # occupied: 傳回 所挑選的位置
    # y_min:     梁筋跑道位置起點
    # y_max:     梁筋跑道位置終點

    f1MAX=1.0   #   均佈加權值
    f2MAX=2.0   #   頭尾加權值:大梁兩側必須要有主筋
    f3MAX=1.0   #   直通跑道加權值：考慮直通優先於彎鉤錨碇

    n_start=y.index(y_min)
    n_stop=y.index(y_max)
    _n_runway=n_stop-n_start+1
    f1=[0]*_n_ab
    f2=[0]*_n_ab
    f3=[0]*_n_ab

    if demand>_n_runway:        # 如果需求超過可以提供的跑道數，
        demand=_n_runway
    _n_space=demand-2+1         # 等距分佈間格數
    
    _pitch=(y_max-y_min)/_n_space   # pitch=總長/間隔數

    #   
    #   均佈加權值
    #
    #   定義加權值函數 w1v : 根據跑道與目標位置的距離決定
    _j=0
    n_wave=demand*2-1   #   上下坡段數
    w1x=[0]*(n_wave)
    w1v=[0]*(n_wave)
    for _ in range(demand-1):
        w1x[_j]=y_min+_pitch*_
        w1v[_j]=f1MAX
        _j+=1
        w1x[_j]=y_min+_pitch*(_+0.5)
        w1v[_j]=f1MAX/3
        _j+=1
    w1x[_j]=y_max
    w1v[_j]=1.0

    for i_wave in range(n_wave-1):
        w1a=w1x[i_wave]
        w1b=w1x[i_wave+1]
        v1a=w1v[i_wave]
        v1b=w1v[i_wave+1]
        for j_runway in range(n_start,n_stop+1):
            if y[j_runway]>=w1a and y[j_runway]<=w1b:
                f1[j_runway]=w1v[i_wave]+(y[j_runway]-w1a)*(w1v[i_wave+1]-w1v[i_wave])/(w1b-w1a)    #   內插求值
                if v1b>v1a:
                    f1[j_runway]+=0.1
                f1[j_runway]=math.sqrt(f1[j_runway])
    #            
    #   頭尾加權值
    #
    f2[n_start]=f2MAX
    f2[n_stop]=f2MAX

    #
    #   直通跑道加權值
    #
    for j_runway in range(n_start,n_stop+1):
        for _, _element in enumerate(out1):
            x=_element['x']
            if abs(y[j_runway]-x)<=1.0:
                _symbol=_element['symbol']
                if _symbol=="═══════":
                    f3[j_runway]=f3MAX
                    break
                break
    f123=[]

    for j in range(n_start,n_stop+1):
        f_value=f1[j]+f2[j]+f3[j]
        f123=f123+[{'runway':j,'value':f_value}]
    
    f123.sort(key=getvalue,reverse=True) # 這一步很重要，按照座標重新排序

    for _ in range(demand):
        j=f123[_]['runway']
        occupied[j]=R_VALUE

    return occupied



######################################
#
#   START main program
#
######################################
os.chdir('/Users/jason/OneDrive/Documents/python/rebar/work')
# C:\Users\jason\OneDrive\Documents\python\rebar\excel-read.xlsx
wb = openpyxl.load_workbook('excel-read.xlsx',data_only=True)     # 開啟 Excel 檔案
names = wb.sheetnames    # 讀取 Excel 裡所有工作表名稱

working_sheet = wb['輸入用']
n_data=working_sheet.cell(1,2).value # read number of data raws
n_data=working_sheet['B1'].value
for i_data in range(n_data):
    i_cell=i_data+3
    widc=working_sheet.cell(i_cell,1).value     #   柱寬
    ba=working_sheet.cell(i_cell,2).value       #   梁寬A
    bb=working_sheet.cell(i_cell,3).value       #   梁寬B
    nc=working_sheet.cell(i_cell,4).value       #   柱主筋數
    shift_a=working_sheet.cell(i_cell,5).value  #   梁偏移A
    shift_b=working_sheet.cell(i_cell,6).value  #   梁偏移B
    demand_a=working_sheet.cell(i_cell,7).value #   梁筋數A
    demand_b=working_sheet.cell(i_cell,8).value #   梁筋數B
    u_layout=working_sheet.cell(i_cell,9).value #   柱筋排列方式='角落' / '均佈' 內定值=角落
    if widc==0 or ( ba==0 and bb==0 ):
        sys.exit('柱寬或梁寬有誤')
    if nc==0:
        sys.exit('excel 輸入錯誤，柱主筋不可為0')
    if not u_layout in ['均佈' ,'角落']:
        sys.exit('excel 輸入錯誤，必須是均佈or角落')

    # 找出可以擺放梁筋各空格的直通或錨碇狀態，放入out1中，這時候還未真正擺放梁筋
    # 定出各跑道的位置與特性：給柱筋、直通梁筋、錨碇梁筋
    # solve_joint(柱寬,梁寬a,梁寬b,柱筋數,梁偏位a,梁偏位b,layout='均佈'或'角落')

    out1=check_rebar_2.build_runway(widc,ba,bb,nc,shift_a,shift_b,u_layout)

    max_wnd=widc*1.0
    EDGE_BM=check_rebar_2.EDGE_DIST_BM
    fig = plt.figure(figsize=[7.,7.])
    plt.axis('off')
    plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei'] 
    plt.rcParams['axes.unicode_minus'] = False

    str1=str(i_data+1)+': '+str(widc)+' '+str(ba)+' '+str(bb)+' 柱筋:'+str(nc)\
        +' 梁筋:'+str(demand_a)+' '+str(demand_b)
    plt.text(0.5,1.05,str1,size=12,ha='center')

    yab=[]
    n_ab=0      #   梁筋全部可以用的空格數
    ya_min=widc # a梁可以用的空格開始位置
    ya_max=0    # a梁可以用的空格結束位置
    yb_min=widc # b梁可以用的空格開始位置
    yb_max=0    # b梁可以用的空格結束位置


    out_x=[]    #   儲存 out1的座標
    for _, i_out1 in enumerate(out1):
        xx=i_out1['x']
        xx_over_w=xx/max_wnd
        symbol=i_out1['symbol']
        out_x=out_x+[xx]

        if symbol=="═══════":
            draw_rebar(xx_over_w,'=')
            ya_min=min(ya_min,xx)
            ya_max=max(ya_max,xx)
            yb_min=min(yb_min,xx)
            yb_max=max(yb_max,xx)
            yab=yab+[xx]
            n_ab+=1
        elif symbol=="●     ●":
            draw_circle(0.05,xx_over_w,32/max_wnd/2,'b')
            draw_circle(0.95,xx_over_w,32/max_wnd/2,'b')
        elif symbol=="══════╝":
            draw_rebar(xx_over_w,'R')
            ya_min=min(ya_min,xx)
            ya_max=max(ya_max,xx)
            yab=yab+[xx]
            n_ab+=1
        elif symbol=="╚══════":
            draw_rebar(xx_over_w,'L')
            yb_min=min(yb_min,xx)
            yb_max=max(yb_max,xx)
            yab=yab+[xx]
            n_ab+=1


    #
    #   依照需求梁筋數目，挑選可用的跑道
    #
    REBAR_LEFT=1
    REBAR_RIGHT=10    
    occupied_a=[0]*n_ab
    occupied_a=select_runway(n_ab,demand_a,yab,occupied_a,out1,ya_min,ya_max,1)
    occupied_b=[0]*n_ab     
    occupied_b=select_runway(n_ab,demand_b,yab,occupied_b,out1,yb_min,yb_max,1)
    #
    #   找出在直通跑道單根交錯的數目 n_single
    #
    n_single=0
    for i in range(n_ab):
        yabi=yab[i]
        i_out=out_x.index(yabi)
        if out1[i_out]['symbol']=='═══════':
            if occupied_a[i]+occupied_b[i]==1:  # 找到單側有鋼筋
                n_single+=1
                if n_single>0 and (n_single % 2)==0:    # merge
                    if yabi in [ya_min,ya_max,yb_min,yb_max]:     #   a梁遇到頭尾跑道
                        occupied_a[i]=1
                        occupied_b[i]=1
                        occupied_a[j]=0
                        occupied_b[j]=0
                    elif yabj in [ya_min,ya_max,yb_min,yb_max]:   #   b梁遇到頭尾跑道
                        occupied_a[i]=0
                        occupied_b[i]=0
                        occupied_a[j]=1
                        occupied_b[j]=1
                    else:
                        occupied_a[i]=1
                        occupied_b[i]=1
                        occupied_a[j]=0
                        occupied_b[j]=0

                j=i                                             #   記住落單的位置，以備後續使用
                yabj=yab[j]

    for j in range(n_ab):
        stat_j=0        
        if occupied_a[j]>=1:
            stat_j+=REBAR_LEFT   # a梁有鋼筋，填 1
        if occupied_b[j]>=1:
            stat_j+=REBAR_RIGHT   # b梁有鋼筋，填 10
        if stat_j==REBAR_LEFT:
            line(-0.1,yab[j]/max_wnd,0.9,yab[j]/max_wnd,'g',8)
        elif stat_j==REBAR_RIGHT:
            line(0.1,yab[j]/max_wnd,1.1,yab[j]/max_wnd,'g',8)
        elif stat_j==REBAR_LEFT+REBAR_RIGHT:
            line(-0.1,yab[j]/max_wnd,1.1,yab[j]/max_wnd,'g',8)

    draw_beam_column_joint(widc/max_wnd,ba/max_wnd,bb/max_wnd,shift_a/max_wnd,shift_b/max_wnd,EDGE_BM/max_wnd)

    # plt.show()
    fig.savefig('plot.'+str(i_data+1)+'.png')
    plt.close()
